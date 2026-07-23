"""Deterministic data profiling and Excel report generation.

Usage:
    python data_profiler.py your_data.csv
    python data_profiler.py your_data.csv my_report.xlsx

The workbook contains Overview, Basic Profile, Advanced Profile, and an optional
Correlation Matrix. It intentionally does not calculate a data quality score.
"""
from __future__ import annotations

import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HEADER_FILL = PatternFill("solid", fgColor="6C72CB")
TITLE_FILL = PatternFill("solid", fgColor="2D2A6E")
STRIPE_FILL = PatternFill("solid", fgColor="EEF0FB")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
TITLE_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=16)
SUBTITLE_FONT = Font(color="6C72CB", italic=True, name="Calibri", size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FORMULA_PREFIXES = ("=", "+", "-", "@")


def normalize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Return a copy with stable, unique string column names."""
    if not isinstance(df, pd.DataFrame):
        raise TypeError("Expected a pandas DataFrame.")
    data = df.copy()
    counts: dict[str, int] = {}
    names: list[str] = []
    renames: list[dict[str, str]] = []
    for index, original in enumerate(data.columns, 1):
        base = str(original).strip() or f"unnamed_{index}"
        counts[base] = counts.get(base, 0) + 1
        unique = base if counts[base] == 1 else f"{base}__{counts[base]}"
        names.append(unique)
        if unique != str(original):
            renames.append({"Original": str(original), "Profiled as": unique})
    data.columns = names
    data.attrs["column_renames"] = renames
    return data


def _hashable_value(value: Any) -> Any:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, (dict, list, tuple, set, np.ndarray)):
        try:
            return json.dumps(value, sort_keys=True, default=str)
        except TypeError:
            return repr(value)
    return value


def _analysis_series(series: pd.Series) -> pd.Series:
    if pd.api.types.is_numeric_dtype(series) or pd.api.types.is_datetime64_any_dtype(series):
        return series
    return series.map(_hashable_value)


def duplicate_row_count(df: pd.DataFrame) -> int:
    data = normalize_dataframe(df)
    try:
        return int(data.duplicated().sum())
    except (TypeError, ValueError):
        comparable = data.apply(lambda column: column.map(_hashable_value))
        return int(comparable.duplicated().sum())


def _excel_text(value: str) -> str:
    """Keep source-derived strings from being interpreted as Excel formulas."""
    stripped = value.lstrip()
    return f"'{value}" if stripped.startswith(FORMULA_PREFIXES) else value


def _excel_value(value: Any) -> Any:
    if isinstance(value, np.integer):
        return int(value)
    if isinstance(value, np.floating):
        return float(value)
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, (dict, list, tuple, set, np.ndarray)):
        value = json.dumps(value, default=str)
    if isinstance(value, str):
        return _excel_text(value)
    return value


def style_header_row(ws, row: int, ncols: int) -> None:
    for column in range(1, ncols + 1):
        cell = ws.cell(row=row, column=column)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def write_df(ws, df: pd.DataFrame, start_row: int = 1, stripe: bool = True) -> int:
    if df.empty and len(df.columns) == 0:
        ws.cell(row=start_row, column=1, value="No profile rows")
        return start_row
    for j, column in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=j, value=_excel_value(str(column)))
    style_header_row(ws, start_row, max(1, len(df.columns)))
    for i, row in enumerate(df.itertuples(index=False), start_row + 1):
        for j, value in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=_excel_value(value))
            cell.border = BORDER
            if stripe and (i - start_row) % 2 == 0:
                cell.fill = STRIPE_FILL
    return start_row + len(df)


def autofit(ws, max_width: int = 42) -> None:
    from openpyxl.utils import get_column_letter
    for index in range(1, ws.max_column + 1):
        length = 8
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=index).value
            if value is not None:
                length = max(length, len(str(value)))
        ws.column_dimensions[get_column_letter(index)].width = min(length + 2, max_width)


def add_title(ws, title: str, subtitle: str, ncols: int) -> None:
    ncols = max(1, ncols)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=1, value=_excel_value(title))
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    sub = ws.cell(row=2, column=1, value=_excel_value(subtitle))
    sub.font = SUBTITLE_FONT
    sub.alignment = Alignment(horizontal="left", indent=1)


def basic_profile(df: pd.DataFrame) -> pd.DataFrame:
    data = normalize_dataframe(df)
    rows: list[dict[str, Any]] = []
    for column in data.columns:
        original = data[column]
        series = _analysis_series(original)
        is_numeric = pd.api.types.is_numeric_dtype(original)
        numeric = pd.to_numeric(original, errors="coerce").replace([np.inf, -np.inf], np.nan) if is_numeric else None
        try:
            unique = int(series.nunique(dropna=True))
        except (TypeError, ValueError):
            series = series.astype(str)
            unique = int(series.nunique(dropna=True))
        try:
            mode = series.mode(dropna=True)
        except (TypeError, ValueError):
            mode = pd.Series(dtype="object")
        try:
            counts = series.value_counts(dropna=True)
            top_frequency = int(counts.iloc[0]) if len(counts) else 0
            top_value = str(counts.index[0]) if len(counts) else None
        except (TypeError, ValueError):
            top_frequency = 0
            top_value = None
        non_null_strings = series.dropna().astype(str)
        rows.append({
            "Column": str(column),
            "Dtype": str(original.dtype),
            "Count": int(original.count()),
            "Missing": int(original.isna().sum()),
            "Missing %": round(float(original.isna().mean() * 100), 2),
            "Unique": unique,
            "Unique %": round(float(unique / len(original) * 100), 2) if len(original) else 0,
            "Top Value": top_value if top_value is not None else (str(mode.iloc[0]) if not mode.empty else None),
            "Top Freq": top_frequency,
            "Mean": round(float(numeric.mean()), 2) if is_numeric and numeric is not None and numeric.count() else None,
            "Std": round(float(numeric.std()), 2) if is_numeric and numeric is not None and numeric.count() > 1 else None,
            "Min": float(numeric.min()) if is_numeric and numeric is not None and numeric.count() else (non_null_strings.min() if len(non_null_strings) else None),
            "Max": float(numeric.max()) if is_numeric and numeric is not None and numeric.count() else (non_null_strings.max() if len(non_null_strings) else None),
        })
    return pd.DataFrame(rows)


def advanced_profile(df: pd.DataFrame) -> pd.DataFrame:
    data = normalize_dataframe(df)
    rows: list[dict[str, Any]] = []
    for column in data.columns:
        original = data[column]
        series = _analysis_series(original)
        is_numeric = pd.api.types.is_numeric_dtype(original)
        outliers = skew = kurt = None
        if is_numeric:
            numeric = pd.to_numeric(original, errors="coerce").replace([np.inf, -np.inf], np.nan)
            if numeric.count() > 1:
                q1, q3 = numeric.quantile([0.25, 0.75])
                iqr = q3 - q1
                outliers = int(((numeric < q1 - 1.5 * iqr) | (numeric > q3 + 1.5 * iqr)).sum())
                skew_value = numeric.skew()
                kurt_value = numeric.kurt()
                skew = round(float(skew_value), 2) if pd.notna(skew_value) else None
                kurt = round(float(kurt_value), 2) if pd.notna(kurt_value) else None
        try:
            unique = int(series.nunique(dropna=True))
        except (TypeError, ValueError):
            unique = int(series.astype(str).nunique(dropna=True))
        cardinality_ratio = round(float(unique / len(series)), 3) if len(series) else 0
        if unique <= 1:
            key_flag = "Constant"
        elif cardinality_ratio > 0.95:
            key_flag = "Likely Key"
        else:
            key_flag = "Categorical/Other"
        pattern_percent = None
        if not is_numeric and series.count():
            patterns = series.dropna().astype(str).apply(lambda value: re.sub(r"[A-Za-z]", "A", re.sub(r"\d", "9", value)))
            frequencies = patterns.value_counts(normalize=True)
            pattern_percent = round(float(frequencies.iloc[0]) * 100, 1) if len(frequencies) else None
        rows.append({
            "Column": str(column),
            "Outlier Count (IQR)": outliers,
            "Skewness": skew,
            "Kurtosis": kurt,
            "Cardinality Ratio": cardinality_ratio,
            "Key Candidate Flag": key_flag,
            "Dominant Pattern %": pattern_percent,
        })
    return pd.DataFrame(rows)


def correlation_matrix(df: pd.DataFrame) -> pd.DataFrame | None:
    data = normalize_dataframe(df)
    numeric = data.select_dtypes(include=np.number).replace([np.inf, -np.inf], np.nan)
    if numeric.shape[1] < 2:
        return None
    correlation = numeric.corr().round(2)
    correlation.insert(0, "Column", correlation.index.astype(str))
    return correlation.reset_index(drop=True)


def build_report(df: pd.DataFrame, source_name: str, out_path: str) -> str:
    data = normalize_dataframe(df)
    workbook = Workbook()
    overview_sheet = workbook.active
    overview_sheet.title = "Overview"
    add_title(overview_sheet, "Data Profiling Report", f"{source_name}  •  generated {datetime.now():%Y-%m-%d %H:%M}", 2)
    overview = pd.DataFrame({
        "Metric": [
            "Rows", "Columns", "Memory Usage (MB)", "Duplicate Rows", "Total Missing Cells",
            "Overall Missing %", "Numeric Columns", "Other Columns", "Renamed Duplicate/Blank Columns", "Generated With",
        ],
        "Value": [
            len(data), data.shape[1], round(float(data.memory_usage(deep=True).sum() / 1_000_000), 2),
            duplicate_row_count(data), int(data.isnull().sum().sum()),
            round(float(data.isnull().sum().sum() / data.size * 100), 2) if data.size else 0,
            data.select_dtypes(include=np.number).shape[1], data.select_dtypes(exclude=np.number).shape[1],
            len(data.attrs.get("column_renames", [])), "Data Profiling Manager by Aanchal Dahiya",
        ],
    })
    write_df(overview_sheet, overview, start_row=4)
    autofit(overview_sheet)

    basic = basic_profile(data)
    basic_sheet = workbook.create_sheet("Basic Profile")
    add_title(basic_sheet, "Basic Profile", "Completeness, uniqueness, frequency, and range", max(1, len(basic.columns)))
    write_df(basic_sheet, basic, start_row=4)
    autofit(basic_sheet)
    if len(basic):
        missing_column = basic.columns.get_loc("Missing %") + 1
        missing_range = f"{basic_sheet.cell(row=5, column=missing_column).coordinate}:{basic_sheet.cell(row=4 + len(basic), column=missing_column).coordinate}"
        basic_sheet.conditional_formatting.add(missing_range, ColorScaleRule(start_type="min", start_color="63BE7B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="F8696B"))

    advanced = advanced_profile(data)
    advanced_sheet = workbook.create_sheet("Advanced Profile")
    add_title(advanced_sheet, "Advanced Profile", "Outliers, distribution shape, key candidates, and patterns", max(1, len(advanced.columns)))
    write_df(advanced_sheet, advanced, start_row=4)
    autofit(advanced_sheet)
    if len(advanced):
        outlier_column = advanced.columns.get_loc("Outlier Count (IQR)") + 1
        outlier_range = f"{advanced_sheet.cell(row=5, column=outlier_column).coordinate}:{advanced_sheet.cell(row=4 + len(advanced), column=outlier_column).coordinate}"
        advanced_sheet.conditional_formatting.add(outlier_range, ColorScaleRule(start_type="min", start_color="63BE7B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="F8696B"))

    correlation = correlation_matrix(data)
    if correlation is not None:
        correlation_sheet = workbook.create_sheet("Correlation Matrix")
        add_title(correlation_sheet, "Correlation Matrix", "Numeric columns only, Pearson correlation", len(correlation.columns))
        write_df(correlation_sheet, correlation, start_row=4, stripe=False)
        autofit(correlation_sheet)
        last_column = len(correlation.columns)
        correlation_range = f"B5:{correlation_sheet.cell(row=4 + len(correlation), column=last_column).coordinate}"
        correlation_sheet.conditional_formatting.add(correlation_range, ColorScaleRule(start_type="num", start_value=-1, start_color="F8696B", mid_type="num", mid_value=0, mid_color="FFFFFF", end_type="num", end_value=1, end_color="63BE7B"))

    if data.attrs.get("column_renames"):
        rename_sheet = workbook.create_sheet("Column Name Notes")
        renames = pd.DataFrame(data.attrs["column_renames"])
        add_title(rename_sheet, "Column Name Notes", "Blank or duplicate column names were made unique for profiling", len(renames.columns))
        write_df(rename_sheet, renames, start_row=4)
        autofit(rename_sheet)

    workbook.save(out_path)
    return out_path


def read_table(path: str) -> pd.DataFrame:
    suffix = Path(path.split("?", 1)[0]).suffix.lower()
    if suffix == ".csv":
        try:
            data = pd.read_csv(path, encoding="utf-8-sig")
        except UnicodeDecodeError:
            data = pd.read_csv(path, encoding="cp1252")
        if data.shape[1] == 1:
            try:
                detected = pd.read_csv(path, sep=None, engine="python", encoding="utf-8-sig")
                if detected.shape[1] > 1:
                    data = detected
            except Exception:
                pass
        return normalize_dataframe(data)
    if suffix in {".xlsx", ".xls"}:
        return normalize_dataframe(pd.read_excel(path))
    if suffix == ".parquet":
        return normalize_dataframe(pd.read_parquet(path))
    raise ValueError("Supported files are CSV, Excel, and Parquet.")


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: python data_profiler.py your_data.csv [output.xlsx]")
        return 1
    source = sys.argv[1]
    output = sys.argv[2] if len(sys.argv) > 2 else "data_profiling_report.xlsx"
    data = read_table(source)
    build_report(data, source, output)
    print(f"Report saved to {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
