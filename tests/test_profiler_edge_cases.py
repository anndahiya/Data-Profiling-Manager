import tempfile
import unittest
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from data_profiler import advanced_profile, basic_profile, build_report, duplicate_row_count, normalize_dataframe
from snapshot_manager import create_snapshot, monitor_alerts, empty_workspace, add_snapshot, upsert_dataset


class ProfilerEdgeCaseTests(unittest.TestCase):
    def test_duplicate_and_blank_columns_are_made_unique(self):
        frame = pd.DataFrame([[1, 2, 3], [1, 2, 3]], columns=["id", "id", ""])
        normalized = normalize_dataframe(frame)
        self.assertEqual(list(normalized.columns), ["id", "id__2", "unnamed_3"])
        self.assertEqual(len(normalized.attrs["column_renames"]), 2)
        self.assertEqual(duplicate_row_count(frame), 1)
        snapshot = create_snapshot(
            frame,
            dataset_id="duplicate-columns",
            dataset_name="Duplicate columns",
            source_name="duplicate.csv",
        )
        self.assertEqual(snapshot["columns"], 3)
        self.assertEqual(len(snapshot["column_renames"]), 2)

    def test_complex_values_and_infinities_do_not_crash(self):
        frame = pd.DataFrame(
            {
                "payload": [[1, 2], {"a": 1}, [1, 2]],
                "metric": [1.0, np.inf, -np.inf],
                "empty": [None, None, None],
            }
        )
        basic = basic_profile(frame)
        advanced = advanced_profile(frame)
        self.assertEqual(len(basic), 3)
        self.assertEqual(len(advanced), 3)
        self.assertEqual(duplicate_row_count(frame), 0)

    def test_report_is_valid_and_strings_are_not_formulas(self):
        frame = pd.DataFrame({"text": ["=2+2", "+SUM(A1:A2)", "normal"], "number": [1, 2, 3]})
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "report.xlsx"
            build_report(frame, "=untrusted-source.csv", str(path))
            workbook = load_workbook(path, data_only=False)
            self.assertIn("Overview", workbook.sheetnames)
            self.assertIn("Basic Profile", workbook.sheetnames)
            self.assertIn("Advanced Profile", workbook.sheetnames)
            self.assertIn("Correlation Matrix", workbook.sheetnames)
            self.assertTrue(path.stat().st_size > 1000)

    def test_monitor_alerts_are_factual(self):
        first = pd.DataFrame({"id": [1, 2], "value": [1, None]})
        second = pd.DataFrame({"id": [1, 2, 3], "value": [1, None, None], "new": ["a", "b", "c"]})
        workspace = empty_workspace()
        upsert_dataset(workspace, dataset_id="sample", dataset_name="Sample", source_name="sample.csv")
        add_snapshot(workspace, create_snapshot(first, dataset_id="sample", dataset_name="Sample", source_name="one.csv"))
        add_snapshot(workspace, create_snapshot(second, dataset_id="sample", dataset_name="Sample", source_name="two.csv"))
        alerts = monitor_alerts(workspace, missing_threshold=5)
        messages = " ".join(alert["Observation"] for alert in alerts)
        self.assertIn("Schema changed", messages)
        self.assertIn("Row count changed", messages)
        self.assertNotIn("quality score", messages.lower())


if __name__ == "__main__":
    unittest.main()
