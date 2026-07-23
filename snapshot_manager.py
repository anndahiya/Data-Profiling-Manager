"""Create, retain, compare, trend, monitor, and export profiling snapshots."""
from __future__ import annotations

import json
import re
import uuid
from datetime import datetime, timezone
from io import BytesIO
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from data_profiler import advanced_profile, basic_profile, correlation_matrix, duplicate_row_count, normalize_dataframe

DATA_VERSION = 2
MAX_RUNS_PER_DATASET = 30
MAX_TOTAL_RUNS = 150
MAX_FAILURES = 100
MAX_BROWSER_DATA_BYTES = 4_000_000
FORMULA_PREFIXES = ("=", "+", "-", "@")

HEADER_FILL = PatternFill("solid", fgColor="6C72CB")
TITLE_FILL = PatternFill("solid", fgColor="2D2A6E")
STRIPE_FILL = PatternFill("solid", fgColor="EEF0FB")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
TITLE_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=16)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def empty_workspace() -> dict[str, Any]:
    return {"version": DATA_VERSION, "datasets": [], "runs": [], "failures": [], "updated_at": None}


def clean_id(value: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9_-]+", "-", value.strip()).strip("-").lower()
    return cleaned or f"dataset-{uuid.uuid4().hex[:8]}"


def unique_dataset_id(workspace: dict[str, Any], requested_name: str) -> str:
    base = clean_id(requested_name)
    existing = {str(item.get("id")) for item in workspace.get("datasets", [])}
    if base not in existing:
        return base
    index = 2
    while f"{base}-{index}" in existing:
        index += 1
    return f"{base}-{index}"


def _safe(value: Any) -> Any:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, np.integer):
        return int(value)
    if isinstance(value, np.floating):
        return float(value)
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _excel_safe(value: Any) -> Any:
    value = _safe(value)
    if isinstance(value, str) and value.lstrip().startswith(FORMULA_PREFIXES):
        return f"'{value}"
    return value


def _records(frame: pd.DataFrame | None) -> list[dict[str, Any]]:
    if frame is None or frame.empty:
        return []
    return [{str(key): _safe(value) for key, value in row.items()} for row in frame.to_dict(orient="records")]


def create_snapshot(
    df: pd.DataFrame,
    *,
    dataset_id: str,
    dataset_name: str,
    source_name: str,
    owner: str = "",
) -> dict[str, Any]:
    data = normalize_dataframe(df)
    basic = basic_profile(data)[
        ["Column", "Dtype", "Count", "Missing", "Missing %", "Unique", "Unique %", "Top Freq", "Mean", "Std"]
    ]
    advanced = advanced_profile(data)
    correlation = correlation_matrix(data)
    now = datetime.now(timezone.utc).isoformat(timespec="seconds")
    return {
        "run_id": uuid.uuid4().hex,
        "status": "success",
        "dataset_id": dataset_id,
        "dataset_name": dataset_name,
        "source_name": source_name,
        "owner": owner,
        "profiled_at": now,
        "rows": int(len(data)),
        "columns": int(data.shape[1]),
        "duplicate_rows": duplicate_row_count(data),
        "missing_cells": int(data.isna().sum().sum()),
        "overall_missing_percent": round(float(data.isna().sum().sum() / data.size * 100), 2) if data.size else 0.0,
        "memory_mb": round(float(data.memory_usage(deep=True).sum() / 1_000_000), 2),
        "numeric_columns": int(data.select_dtypes(include=np.number).shape[1]),
        "categorical_columns": int(data.select_dtypes(exclude=np.number).shape[1]),
        "basic_profile": _records(basic),
        "advanced_profile": _records(advanced),
        "correlation_profile": _records(correlation),
        "schema": {str(column): str(dtype) for column, dtype in data.dtypes.items()},
        "column_renames": data.attrs.get("column_renames", []),
        "ai_summary": None,
    }


def add_failure(
    workspace: dict[str, Any],
    *,
    dataset_name: str,
    source_name: str,
    error_message: str,
    dataset_id: str | None = None,
) -> dict[str, Any]:
    failure = {
        "event_id": uuid.uuid4().hex,
        "status": "failed",
        "dataset_id": dataset_id,
        "dataset_name": dataset_name or "Unregistered dataset",
        "source_name": source_name,
        "profiled_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "error": error_message[:500],
    }
    failures = workspace.setdefault("failures", [])
    failures.append(failure)
    failures[:] = failures[-MAX_FAILURES:]
    workspace["updated_at"] = datetime.now(timezone.utc).isoformat(timespec="seconds")
    return failure


def upsert_dataset(
    workspace: dict[str, Any],
    *,
    dataset_id: str,
    dataset_name: str,
    source_name: str,
    owner: str = "",
    description: str = "",
    tags: list[str] | None = None,
) -> None:
    datasets = workspace.setdefault("datasets", [])
    existing = next((item for item in datasets if item.get("id") == dataset_id), None)
    now = datetime.now(timezone.utc).isoformat(timespec="seconds")
    payload = {
        "id": dataset_id,
        "name": dataset_name,
        "source_name": source_name,
        "owner": owner,
        "description": description,
        "tags": tags or [],
        "updated_at": now,
    }
    if existing:
        schedule = existing.get("schedule")
        existing.update(payload)
        if schedule:
            existing["schedule"] = schedule
    else:
        payload["created_at"] = now
        datasets.append(payload)
    datasets.sort(key=lambda item: str(item.get("name", "")).lower())


def successful_runs(workspace: dict[str, Any]) -> list[dict[str, Any]]:
    return [run for run in workspace.get("runs", []) if run.get("status", "success") == "success"]


def add_snapshot(workspace: dict[str, Any], snapshot: dict[str, Any]) -> None:
    snapshot.setdefault("status", "success")
    runs = workspace.setdefault("runs", [])
    if any(run.get("run_id") == snapshot.get("run_id") for run in runs):
        return
    runs.append(snapshot)
    runs.sort(key=lambda item: item.get("profiled_at", ""))
    dataset_id = snapshot.get("dataset_id")
    matching = [run for run in runs if run.get("dataset_id") == dataset_id]
    if len(matching) > MAX_RUNS_PER_DATASET:
        remove_ids = {run["run_id"] for run in matching[:-MAX_RUNS_PER_DATASET]}
        runs[:] = [run for run in runs if run.get("run_id") not in remove_ids]
    if len(runs) > MAX_TOTAL_RUNS:
        runs[:] = runs[-MAX_TOTAL_RUNS:]
    while len(json.dumps(workspace, ensure_ascii=False, default=str).encode("utf-8")) > MAX_BROWSER_DATA_BYTES and len(runs) > 1:
        runs.pop(0)
    workspace["updated_at"] = datetime.now(timezone.utc).isoformat(timespec="seconds")


def dataset_runs(workspace: dict[str, Any], dataset_id: str) -> list[dict[str, Any]]:
    return sorted(
        [run for run in successful_runs(workspace) if run.get("dataset_id") == dataset_id],
        key=lambda item: item.get("profiled_at", ""),
    )


def latest_run(workspace: dict[str, Any], dataset_id: str | None = None) -> dict[str, Any] | None:
    runs = successful_runs(workspace)
    if dataset_id:
        runs = [run for run in runs if run.get("dataset_id") == dataset_id]
    return max(runs, key=lambda item: item.get("profiled_at", ""), default=None)


def find_run(workspace: dict[str, Any], run_id: str | None) -> dict[str, Any] | None:
    if not run_id:
        return None
    return next((run for run in successful_runs(workspace) if run.get("run_id") == run_id), None)


def compare_runs(older: dict[str, Any], newer: dict[str, Any]) -> dict[str, Any]:
    old_schema = older.get("schema", {})
    new_schema = newer.get("schema", {})
    added = sorted(set(new_schema) - set(old_schema))
    removed = sorted(set(old_schema) - set(new_schema))
    dtype_changes = [
        {"Column": column, "Before": old_schema[column], "After": new_schema[column]}
        for column in sorted(set(old_schema) & set(new_schema))
        if old_schema[column] != new_schema[column]
    ]
    old_basic = {row.get("Column"): row for row in older.get("basic_profile", [])}
    new_basic = {row.get("Column"): row for row in newer.get("basic_profile", [])}
    column_changes: list[dict[str, Any]] = []
    for column in sorted(set(old_basic) & set(new_basic)):
        old_row, new_row = old_basic[column], new_basic[column]
        old_missing = float(old_row.get("Missing %") or 0)
        new_missing = float(new_row.get("Missing %") or 0)
        old_unique = int(old_row.get("Unique") or 0)
        new_unique = int(new_row.get("Unique") or 0)
        if old_missing != new_missing or old_unique != new_unique:
            column_changes.append(
                {
                    "Column": column,
                    "Missing % before": old_missing,
                    "Missing % after": new_missing,
                    "Missing % change": round(new_missing - old_missing, 2),
                    "Unique before": old_unique,
                    "Unique after": new_unique,
                    "Unique change": new_unique - old_unique,
                }
            )
    return {
        "summary": {
            "Rows": int(newer.get("rows", 0)) - int(older.get("rows", 0)),
            "Columns": int(newer.get("columns", 0)) - int(older.get("columns", 0)),
            "Duplicate rows": int(newer.get("duplicate_rows", 0)) - int(older.get("duplicate_rows", 0)),
            "Missing cells": int(newer.get("missing_cells", 0)) - int(older.get("missing_cells", 0)),
            "Overall missing %": round(float(newer.get("overall_missing_percent", 0)) - float(older.get("overall_missing_percent", 0)), 2),
            "Memory MB": round(float(newer.get("memory_mb", 0)) - float(older.get("memory_mb", 0)), 2),
        },
        "added_columns": added,
        "removed_columns": removed,
        "dtype_changes": dtype_changes,
        "column_changes": column_changes,
    }


def trend_frame(runs: list[dict[str, Any]]) -> pd.DataFrame:
    rows = [
        {
            "Profiled at": pd.to_datetime(run.get("profiled_at"), utc=True, errors="coerce"),
            "Rows": int(run.get("rows", 0)),
            "Columns": int(run.get("columns", 0)),
            "Duplicate rows": int(run.get("duplicate_rows", 0)),
            "Missing cells": int(run.get("missing_cells", 0)),
            "Overall missing %": float(run.get("overall_missing_percent", 0)),
            "Memory MB": float(run.get("memory_mb", 0)),
        }
        for run in runs
    ]
    frame = pd.DataFrame(rows)
    if frame.empty:
        return frame
    return frame.dropna(subset=["Profiled at"]).sort_values("Profiled at")


def update_ai_summary(workspace: dict[str, Any], run_id: str, summary: str) -> None:
    run = find_run(workspace, run_id)
    if run:
        run["ai_summary"] = summary
        workspace["updated_at"] = datetime.now(timezone.utc).isoformat(timespec="seconds")


def monitor_alerts(workspace: dict[str, Any], missing_threshold: float = 5.0) -> list[dict[str, Any]]:
    alerts: list[dict[str, Any]] = []
    for dataset in workspace.get("datasets", []):
        runs = dataset_runs(workspace, dataset.get("id", ""))
        if not runs:
            alerts.append({"Severity": "Info", "Dataset": dataset.get("name"), "Observation": "Registered but never profiled", "Destination": "Profile"})
            continue
        current = runs[-1]
        if float(current.get("overall_missing_percent", 0)) >= missing_threshold:
            alerts.append({"Severity": "Review", "Dataset": dataset.get("name"), "Observation": f"Overall missing is {float(current.get('overall_missing_percent', 0)):.2f}% (threshold {missing_threshold:.0f}%)", "Destination": "Report viewer"})
        if int(current.get("duplicate_rows", 0)) > 0:
            alerts.append({"Severity": "Review", "Dataset": dataset.get("name"), "Observation": f"{int(current.get('duplicate_rows', 0)):,} duplicate rows detected", "Destination": "Report viewer"})
        if len(runs) >= 2:
            changes = compare_runs(runs[-2], runs[-1])
            if changes["added_columns"] or changes["removed_columns"] or changes["dtype_changes"]:
                alerts.append({"Severity": "Change", "Dataset": dataset.get("name"), "Observation": "Schema changed since the previous run", "Destination": "Compare"})
            if changes["summary"]["Rows"] != 0:
                alerts.append({"Severity": "Change", "Dataset": dataset.get("name"), "Observation": f"Row count changed by {changes['summary']['Rows']:+,}", "Destination": "Compare"})
    return alerts


def workspace_to_json(workspace: dict[str, Any]) -> str:
    return json.dumps(workspace, ensure_ascii=False, separators=(",", ":"), default=str)


def workspace_from_json(value: str) -> dict[str, Any]:
    if not value:
        return empty_workspace()
    parsed = json.loads(value)
    if not isinstance(parsed, dict):
        raise ValueError("Backup must be a JSON object.")
    parsed.setdefault("version", DATA_VERSION)
    parsed.setdefault("datasets", [])
    parsed.setdefault("runs", [])
    parsed.setdefault("failures", [])
    for run in parsed["runs"]:
        run.setdefault("status", "success")
        run.setdefault("correlation_profile", [])
        run.setdefault("column_renames", [])
    parsed["version"] = DATA_VERSION
    return parsed


def _write_table(ws, rows: list[dict[str, Any]], start_row: int = 4) -> None:
    if not rows:
        ws.cell(row=start_row, column=1, value="No saved rows")
        return
    columns = list(rows[0].keys())
    for index, column in enumerate(columns, 1):
        cell = ws.cell(row=start_row, column=index, value=_excel_safe(str(column)))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    for row_index, row in enumerate(rows, start_row + 1):
        for column_index, column in enumerate(columns, 1):
            cell = ws.cell(row=row_index, column=column_index, value=_excel_safe(row.get(column)))
            cell.border = BORDER
            if (row_index - start_row) % 2 == 0:
                cell.fill = STRIPE_FILL
    from openpyxl.utils import get_column_letter
    for index in range(1, ws.max_column + 1):
        width = max(
            (len(str(ws.cell(row=row, column=index).value)) for row in range(1, ws.max_row + 1) if ws.cell(row=row, column=index).value is not None),
            default=8,
        )
        ws.column_dimensions[get_column_letter(index)].width = min(width + 2, 42)


def _title(ws, title: str, subtitle: str, ncols: int) -> None:
    ncols = max(1, ncols)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=1, value=_excel_safe(title))
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left", indent=1)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.cell(row=2, column=1, value=_excel_safe(subtitle))


def snapshot_report_bytes(snapshot: dict[str, Any]) -> bytes:
    workbook = Workbook()
    overview_sheet = workbook.active
    overview_sheet.title = "Overview"
    overview_rows = [
        {"Metric": "Dataset", "Value": snapshot.get("dataset_name")},
        {"Metric": "Source", "Value": snapshot.get("source_name")},
        {"Metric": "Profiled at", "Value": snapshot.get("profiled_at")},
        {"Metric": "Rows", "Value": snapshot.get("rows")},
        {"Metric": "Columns", "Value": snapshot.get("columns")},
        {"Metric": "Duplicate rows", "Value": snapshot.get("duplicate_rows")},
        {"Metric": "Missing cells", "Value": snapshot.get("missing_cells")},
        {"Metric": "Overall missing %", "Value": snapshot.get("overall_missing_percent")},
        {"Metric": "Memory MB", "Value": snapshot.get("memory_mb")},
        {"Metric": "Generated with", "Value": "Data Profiling Manager by Aanchal Dahiya"},
    ]
    _title(overview_sheet, "Data Profiling Snapshot", "Recreated from saved aggregate profiling metrics", 2)
    _write_table(overview_sheet, overview_rows)
    for sheet_name, title, rows in [
        ("Basic Profile", "Basic Profile", snapshot.get("basic_profile", [])),
        ("Advanced Profile", "Advanced Profile", snapshot.get("advanced_profile", [])),
        ("Correlation Matrix", "Correlation Matrix", snapshot.get("correlation_profile", [])),
    ]:
        if sheet_name == "Correlation Matrix" and not rows:
            continue
        sheet = workbook.create_sheet(sheet_name)
        _title(sheet, title, "Saved profiling snapshot", len(rows[0]) if rows else 1)
        _write_table(sheet, rows)
    if snapshot.get("ai_summary"):
        ai_sheet = workbook.create_sheet("AI Explanation")
        _title(ai_sheet, "Gemini Explanation", "Generated from aggregate profiling metrics", 1)
        ai_sheet.cell(row=4, column=1, value=_excel_safe(snapshot["ai_summary"]))
        ai_sheet.cell(row=4, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ai_sheet.column_dimensions["A"].width = 110
        ai_sheet.row_dimensions[4].height = 300
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
